import streamlit as st
import requests
import pandas as pd
import io
from openpyxl.utils import get_column_letter

CLIENT_ID    = st.secrets["CLIENT_ID"]
CLIENT_SECRET = st.secrets["CLIENT_SECRET"]

TOKEN_URL = "https://entreprise.francetravail.fr/connexion/oauth2/access_token?realm=/partenaire"
API_BASE = "https://api.francetravail.io/partenaire/rome-metiers"
SCOPES = "nomenclatureRome api_rome-metiersv1"

if 'search_done' not in st.session_state:
    st.session_state.search_done = False
    st.session_state.statuts = []
    st.session_state.codes_list = []

def get_token():
    data = {
        "grant_type": "client_credentials",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "scope": SCOPES,
    }
    r = requests.post(TOKEN_URL, data=data)
    r.raise_for_status()
    return r.json()["access_token"]

def get_metier(code_rome):
    token = get_token()
    headers = {"Authorization": f"Bearer {token}"}
    
    champs = (
        "code,"
        "libelle,"
        "contextestravail(categorie,libelle),"
    )
    
    url = f"{API_BASE}/v1/metiers/metier/{code_rome}"
    params = {"champs": champs}
    
    r = requests.get(url, headers=headers, params=params)
    r.raise_for_status()
    return r.json()

def get_contextes_by_categorie(metier, categorie):
    contextes = []
    if 'contextesTravail' in metier:
        for ctx in metier['contextesTravail']:
            if ctx.get('categorie') == categorie:
                libelle = ctx.get('libelle', '').strip()
                if libelle:
                    contextes.append(libelle)
    return contextes

def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
            
        elif isinstance(v, list):
            for i, item in enumerate(v):
                if isinstance(item, dict):
                    items.extend(flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
                else:
                    items.append((f"{new_key}_{i}", item))
        else:
            items.append((new_key, v))
            
    return dict(items)

def is_fipu(conditions_str: str, horaires_str: str) -> bool:
    if not conditions_str and not horaires_str:
        return False
    
    criteres = [
        "En altitude",
        "En milieu nucléaire",
        "En milieu hyperbare",
        "En milieu exigu ou confiné",
        "En grande hauteur",
        "En zone frigorifique",
        "Exposition à de hautes températures",
        "En environnement climatique difficile",
        "Manipulation d'un engin, équipement ou outil dangereux",
        "Port et manipulation de charges lourdes ou encombrantes",
        "Position pénible",
        "Station debout prolongée",
        "Travail répétitif ou cadence imposée",
        "En environnement bruyant",
        "Travail dans des environnements hostiles et dangereux",
        "Exposition à de basses températures",
        "Exposition possible à gaz, aérosol, fumées …",
        "Station assise prolongée",
        "Risques de chutes",
        "Travail dans des milieux difficiles et exigeants pour l'humain",
        "Travail posté (2x8, 3x8, 5x8, etc.)",
        "Travail de nuit",
        "Travail en astreinte",
        "Travail en horaires décalés",
        "Travail par roulement"
    ]
    
    texte = (conditions_str + " " + horaires_str).lower()
    return any(critere.lower() in texte for critere in criteres)

def create_enriched_df(metiers_data):
    rows = []
    
    for metier in metiers_data:
        flat = flatten_dict(metier)
        
        conditions = get_contextes_by_categorie(metier, "CONDITIONS_TRAVAIL")
        horaires   = get_contextes_by_categorie(metier, "HORAIRE_ET_DUREE_TRAVAIL")
        
        conditions_joined = ', '.join(conditions) if conditions else ''
        horaires_joined   = ', '.join(horaires)   if horaires   else ''
        
        flat['Conditions de travail et risques professionnels'] = conditions_joined
        flat['Horaires et durée du travail'] = horaires_joined
        flat['FIPU'] = "OUI" if is_fipu(conditions_joined, horaires_joined) else "NON"
        
        rows.append(flat)
    
    df = pd.DataFrame(rows)
    
    desired_order = [
        'code',
        'libelle',
        'FIPU',
        'Conditions de travail et risques professionnels',
        'Horaires et durée du travail'
    ]
    
    remaining_cols = [c for c in df.columns if c not in desired_order]
    final_order = desired_order
    
    return df[final_order]

st.title("🔎 Recherche Multi-Métiers ROME")
st.markdown("**Entrez plusieurs codes ROME (1 par ligne) et consultez les résultats détaillés**")

codes_input = st.text_area(
    "Codes ROME (un par ligne, ex: A1413\nM1805\nH1203)",
    height=150,
    placeholder="A1413\nM1805\nH1203"
)

if st.button("🔍 Rechercher TOUS les métiers", type="primary"):
    if not codes_input.strip():
        st.warning("⚠️ Veuillez entrer au moins un code ROME.")
    else:
        seen = set()
        codes_list = []
        for line in codes_input.strip().split('\n'):
            code = line.strip().upper()
            if code and code not in seen:
                seen.add(code)
                codes_list.append(code)

        original_len = len([code.strip().upper() for code in codes_input.strip().split('\n') if code.strip()])
        if len(codes_list) < original_len:
            st.info(f"ℹ️ {original_len - len(codes_list)} doublon(s) ignoré(s) (ordre conservé)")
        
        if not codes_list:
            st.warning("⚠️ Aucun code ROME valide détecté.")
        else:
            st.info(f"🔄 Recherche de **{len(codes_list)}** métiers...")
            
            progress_bar = st.progress(0)
            metiers_data = []
            statuts = []
            
            for i, code_rome in enumerate(codes_list):
                try:
                    metier = get_metier(code_rome)
                    libelle = metier.get('libelle', 'Sans libellé')
                    metiers_data.append(metier)
                    statuts.append({
                        'code': code_rome,
                        'libelle': libelle,
                        'metier_data': metier,
                        'success': True
                    })
                except requests.HTTPError:
                    statuts.append({
                        'code': code_rome,
                        'libelle': 'Non trouvé',
                        'success': False
                    })
                except Exception as e:
                    statuts.append({
                        'code': code_rome,
                        'libelle': f'Erreur: {str(e)[:30]}…',
                        'success': False
                    })
                
                progress_bar.progress((i + 1) / len(codes_list))
                
            st.session_state.statuts = statuts
            st.session_state.reussis_data = [s['metier_data'] for s in statuts if s.get('success', False)]
            st.session_state.codes_list = codes_list
            st.session_state.search_done = True


if st.session_state.search_done:
    statuts = st.session_state.statuts
    reussis_data = st.session_state.reussis_data

    st.subheader("📊 Résumé de la recherche")
    
    reussis = sum(1 for s in statuts if s.get('success', False))
    col1, col2 = st.columns([3, 1])
    with col1:
        st.metric("Métiers trouvés", f"{reussis} / {len(st.session_state.codes_list)}")
    
    if reussis_data:
        df = create_enriched_df(reussis_data)
        
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Metiers_ROME', index=False)
            
            worksheet = writer.sheets['Metiers_ROME']
            
            for col_idx, column_cells in enumerate(worksheet.columns, start=1):
                column_letter = get_column_letter(col_idx)
                header_value = worksheet[f"{column_letter}1"].value
                if header_value:
                    length = len(str(header_value)) + 5
                    width = min(length, 80)
                    worksheet.column_dimensions[column_letter].width = width
            
            worksheet.freeze_panes = "A2"
        
        excel_buffer.seek(0)
        
        st.download_button(
            label=f"📥 Télécharger Excel ({len(reussis_data)} métiers)",
            data=excel_buffer.getvalue(),
            file_name=f"ROME_multi_metiers_{len(reussis_data)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    else:
        st.info("Aucun métier trouvé → pas de fichier à télécharger.")
    
    st.divider()

    st.subheader("📋 Détails par métier")
    
    for statut in statuts:
        code_rome = statut['code']
        libelle = statut['libelle']
        
        if statut.get('success', False):
            metier_data = statut['metier_data']
            
            conditions_joined = ', '.join(get_contextes_by_categorie(metier_data, "CONDITIONS_TRAVAIL"))
            horaires_joined   = ', '.join(get_contextes_by_categorie(metier_data, "HORAIRE_ET_DUREE_TRAVAIL"))
            
            fipu_oui = is_fipu(conditions_joined, horaires_joined)
            
            st.success(f"✅ **{libelle}** ({code_rome})")
            
            if fipu_oui:
                st.success("**FIPU : OUI** ✅")
            else:
                st.error("**FIPU : NON** ❌")
            
            st.markdown("**🏭 Conditions de travail et risques professionnels :**")
            if conditions_joined:
                for item in conditions_joined.split(', '):
                    st.markdown(f"- {item}")
            else:
                st.markdown("*Aucune condition trouvée*")
            
            st.markdown("**⏰ Horaires et durée du travail :**")
            if horaires_joined:
                for item in horaires_joined.split(', '):
                    st.markdown(f"- {item}")
            else:
                st.markdown("*Aucun horaire spécifique trouvé*")
            
            st.divider()
        else:
            st.error(f"❌ **{code_rome}** - {libelle}")
            st.divider()

    reussis_data = [s['metier_data'] for s in statuts if s.get('success', False)]
    if reussis_data:
        df = create_enriched_df(reussis_data)
        
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Metiers_ROME', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Metiers_ROME']
            
            for col_idx, column_cells in enumerate(worksheet.columns, start=1):
                column_letter = get_column_letter(col_idx)
                header_value = worksheet[f"{column_letter}1"].value
                
                if header_value:
                    length = len(str(header_value)) + 5
                    width = min(length, 80)
                    worksheet.column_dimensions[column_letter].width = width
            
            worksheet.freeze_panes = "A2"
        
        excel_buffer.seek(0)
        
        st.download_button(
            label=f"📊 Télécharger Excel ({len(reussis_data)} métiers)",
            data=excel_buffer.getvalue(),
            file_name=f"ROME_multi_metiers_{len(reussis_data)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with st.expander("💡 Exemple d'utilisation"):
    st.code("""
A1413
M1805
H1203
K2110
""", language="text")
